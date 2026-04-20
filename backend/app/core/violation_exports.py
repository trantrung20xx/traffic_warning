from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Optional
from urllib.parse import urljoin

from app.core.timezone import ensure_utc_datetime, to_vietnam_datetime

VEHICLE_TYPE_LABELS = {
    "motorcycle": "Xe máy",
    "car": "Ô tô",
    "truck": "Xe tải",
    "bus": "Xe buýt",
}

VIOLATION_LABELS = {
    "wrong_lane": "Đi sai làn",
    "vehicle_type_not_allowed": "Loại phương tiện không đúng quy định",
    "turn_left_not_allowed": "Rẽ trái không đúng quy định",
    "turn_right_not_allowed": "Rẽ phải không đúng quy định",
    "turn_straight_not_allowed": "Đi thẳng không đúng quy định",
    "turn_u_turn_not_allowed": "Quay đầu không đúng quy định",
}

CSV_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("stt", "STT"),
    ("timestamp", "Thời gian vi phạm"),
    ("camera_id", "Camera"),
    ("violation", "Loại vi phạm"),
    ("vehicle_type", "Loại phương tiện"),
    ("vehicle_id", "Vehicle ID"),
    ("lane_id", "Làn"),
    ("location", "Địa điểm / khu vực"),
    ("image_path", "Đường dẫn ảnh vi phạm"),
]

XLSX_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("stt", "STT"),
    ("timestamp", "Thời gian vi phạm"),
    ("camera_id", "Camera"),
    ("violation", "Loại vi phạm"),
    ("vehicle_type", "Loại phương tiện"),
    ("vehicle_id", "Vehicle ID"),
    ("lane_id", "Làn"),
    ("location", "Địa điểm"),
    ("image_path", "Đường dẫn ảnh"),
]


def _parse_iso_datetime(value: Optional[str]) -> Optional[datetime]:
    """Phân tích chuỗi thời gian ISO và chuẩn hóa về UTC nếu có giá trị."""
    if not value:
        return None
    return ensure_utc_datetime(datetime.fromisoformat(value))


def _format_display_timestamp(value: Optional[str]) -> str:
    parsed = _parse_iso_datetime(value)
    if parsed is None:
        return "-"
    return to_vietnam_datetime(parsed).strftime("%d/%m/%Y %H:%M:%S")


def _format_filename_date(value: Optional[str], fallback: str) -> str:
    parsed = _parse_iso_datetime(value)
    if parsed is None:
        return fallback
    return to_vietnam_datetime(parsed).strftime("%Y-%m-%d")


def _get_label(labels: dict[str, str], value: Optional[str]) -> str:
    if value is None or value == "":
        return "-"
    return labels.get(value, value)


def _format_value(value) -> str:
    if value is None or value == "":
        return "-"
    return str(value)


def _build_location_label(location: Optional[dict]) -> str:
    if not location:
        return "-"
    road_name = location.get("road_name")
    intersection = location.get("intersection")
    parts = [part for part in [road_name, intersection] if part]
    return " · ".join(parts) if parts else "-"


def _build_evidence_link(row: dict, *, base_url: str) -> str:
    """Tạo URL tuyệt đối đến ảnh bằng chứng để file export mở được trực tiếp."""
    value = row.get("image_url") or row.get("image_path")
    if not value:
        return "-"
    value = str(value).strip()
    if not value:
        return "-"
    if value.startswith(("http://", "https://")):
        return value
    if value.startswith("/"):
        return urljoin(base_url, value)
    return value


def build_violation_export_rows(rows: list[dict], *, base_url: str) -> list[dict]:
    """Chuẩn hóa dữ liệu lịch sử vi phạm thành các cột phục vụ xuất báo cáo."""
    export_rows = []
    for index, row in enumerate(rows, start=1):
        export_rows.append(
            {
                "stt": str(index),
                "timestamp": _format_display_timestamp(row.get("timestamp")),
                "camera_id": _format_value(row.get("camera_id")),
                "violation": _get_label(VIOLATION_LABELS, row.get("violation")),
                "vehicle_type": _get_label(VEHICLE_TYPE_LABELS, row.get("vehicle_type")),
                "vehicle_id": _format_value(row.get("vehicle_id")),
                "lane_id": _format_value(row.get("lane_id")),
                "location": _build_location_label(row.get("location")),
                "image_path": _build_evidence_link(row, base_url=base_url),
            }
        )
    return export_rows


def build_violation_export_filename(*, extension: str, from_ts: Optional[str], to_ts: Optional[str]) -> str:
    """Sinh tên file export theo khoảng thời gian người dùng đã lọc."""
    start = _format_filename_date(from_ts, "start")
    end = _format_filename_date(to_ts, "end")
    return f"violation_history_{start}_{end}.{extension}"


def build_violation_history_csv(rows: list[dict]) -> bytes:
    """Đóng gói lịch sử vi phạm thành nội dung CSV có BOM để Excel đọc tiếng Việt đúng."""
    buffer = StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in CSV_EXPORT_COLUMNS])
    for row in rows:
        writer.writerow([row[key] for key, _ in CSV_EXPORT_COLUMNS])
    return buffer.getvalue().encode("utf-8-sig")


def build_violation_history_xlsx(rows: list[dict]) -> bytes:
    """Đóng gói lịch sử vi phạm thành tệp Excel với tiêu đề và độ rộng cột dễ đọc."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError("Thiếu thư viện openpyxl. Cài dependency backend để xuất file Excel.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Violation History"
    worksheet.freeze_panes = "A2"

    headers = [label for _, label in XLSX_EXPORT_COLUMNS]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row[key] for key, _ in XLSX_EXPORT_COLUMNS])

    for column_index, (key, _) in enumerate(XLSX_EXPORT_COLUMNS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(str(worksheet[f"{column_letter}1"].value or ""))
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet[f"{column_letter}{row_index}"].value
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
